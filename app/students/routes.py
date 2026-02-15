import csv
import io
import os
from collections import Counter
from datetime import datetime
import html
import tempfile
import re
import unicodedata

from flask import Blueprint, abort, current_app, flash, redirect, render_template, request, send_file, send_from_directory, session, url_for
from flask_login import current_user, login_required
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from app.extensions import csrf, db
from argon2 import PasswordHasher

from app.models import (
    AgencySubscription,
    ArrivalSupport,
    AuditLog,
    Booking,
    Branch,
    CasePayment,
    CaseStage,
    CommissionRecord,
    Document,
    EmailDispatch,
    Guardian,
    InviteToken,
    Student,
    StudentAuth,
    StudentCV,
    StudentDocument,
    StudentDocumentFolder,
    StudyCase,
)
from app.students.forms import GuardianForm, StudentCVForm, StudentDocumentForm, StudentFolderCreateForm, StudentForm
from app.utils.audit import add_audit_log
from app.utils.authz import can_access_branch, is_founder, normalized_role, role_required, scope_query_by_branch, user_branch_ids
from app.utils.files import save_uploaded_file


students_bp = Blueprint("students", __name__, url_prefix="/students")
password_hasher = PasswordHasher()
FOLDER_LABELS = {
    "studentgator": "STUDENTGATOR",
    "student_gator": "STUDENTGATOR",
    "uco": "UCO",
    "cie": "CIE",
    "uit": "UIT",
    "iut": "UIT",
}
FOLDER_ALIASES = {
    "student_gator": "studentgator",
    "studentgator": "studentgator",
    "cie": "cie",
    "uco": "uco",
    "iut": "uit",
    "uit": "uit",
}
CANONICAL_TO_STORED = {
    "studentgator": ["studentgator", "student_gator"],
    "uco": ["uco"],
    "cie": ["cie"],
    "uit": ["uit", "iut"],
}


def generate_matricule(reserved=None):
    reserved = reserved or set()
    year = datetime.utcnow().year
    # Unique key is global on students.matricule, including soft-deleted rows.
    index = Student.query.count() + 1
    while True:
        candidate = f"IF-{year}-{index:05d}"
        if candidate not in reserved and not Student.query.filter(Student.matricule == candidate).first():
            return candidate
        index += 1


def enforce_student_access(student):
    if not can_access_branch(student.branch_id):
        abort(403)


def get_active_student_or_404(student_id):
    return Student.query.filter(Student.id == student_id, Student.deleted_at.is_(None)).first_or_404()


def set_branch_choices(form):
    actor_branch_id = resolve_actor_branch_id()
    if actor_branch_id:
        branch = Branch.query.get(actor_branch_id)
        if branch:
            form.branch_id.choices = [(branch.id, f"{branch.name} ({branch.country_code})")]
            form.branch_id.data = branch.id
            return

    if normalized_role(current_user.role) == "IT":
        choices = [(b.id, f"{b.name} ({b.country_code})") for b in Branch.query.order_by(Branch.name.asc()).all()]
        form.branch_id.choices = choices
        if choices and not form.branch_id.data:
            form.branch_id.data = choices[0][0]
        return

    branch = Branch.query.get(current_user.branch_id) if current_user.branch_id else None
    if branch:
        form.branch_id.choices = [(branch.id, f"{branch.name} ({branch.country_code})")]
        form.branch_id.data = branch.id
    else:
        form.branch_id.choices = []


def resolve_actor_branch_id():
    role = normalized_role(current_user.role)
    if role in ("ADMIN_BRANCH", "EMPLOYEE"):
        return current_user.branch_id
    if role == "FOUNDER":
        if current_user.branch_id:
            return current_user.branch_id
        owner_sub = AgencySubscription.query.filter_by(owner_user_id=current_user.id).first()
        if owner_sub and owner_sub.branch_id:
            current_user.branch_id = owner_sub.branch_id
            db.session.commit()
            return owner_sub.branch_id
    if role == "IT":
        scoped_id = session.get("it_scope_branch_id")
        if scoped_id:
            return scoped_id
    return None


def get_or_create_student_cv(student_id):
    row = StudentCV.query.filter_by(student_id=student_id).first()
    if row is None:
        row = StudentCV(
            student_id=student_id,
            show_hobbies=True,
            show_languages=True,
            show_skills=True,
            show_education=True,
            show_professional_experience=True,
            show_extra_experience=False,
            show_software=True,
            show_social_links=True,
        )
        db.session.add(row)
        db.session.commit()
    return row


def _split_multiline(value):
    if not value:
        return []
    return [line.strip() for line in value.replace("\r", "").split("\n") if line.strip()]


def _compression_profile(target_mb):
    mapping = {
        2: {"jpeg_quality": 45, "max_dim": 1100},
        3: {"jpeg_quality": 55, "max_dim": 1300},
        4: {"jpeg_quality": 65, "max_dim": 1600},
        5: {"jpeg_quality": 75, "max_dim": 1800},
    }
    return mapping.get(target_mb, mapping[4])


def normalize_folder_name(raw):
    value = (raw or "").strip().lower()
    if not value:
        return ""
    value = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode("ascii")
    value = value.replace("-", "_").replace(" ", "_")
    value = re.sub(r"[^a-z0-9_]", "", value)
    value = re.sub(r"_+", "_", value).strip("_")
    if not value:
        return ""
    return FOLDER_ALIASES.get(value, value)


def _purge_student_related_data(student_id):
    case_ids = [row[0] for row in db.session.query(StudyCase.id).filter(StudyCase.student_id == student_id).all()]
    if case_ids:
        CaseStage.query.filter(CaseStage.case_id.in_(case_ids)).delete(synchronize_session=False)
        ArrivalSupport.query.filter(ArrivalSupport.case_id.in_(case_ids)).delete(synchronize_session=False)
        CasePayment.query.filter(CasePayment.case_id.in_(case_ids)).delete(synchronize_session=False)
        CommissionRecord.query.filter(CommissionRecord.case_id.in_(case_ids)).delete(synchronize_session=False)
        Document.query.filter(Document.case_id.in_(case_ids)).delete(synchronize_session=False)
    Document.query.filter_by(student_id=student_id).delete(synchronize_session=False)
    StudyCase.query.filter_by(student_id=student_id).delete(synchronize_session=False)
    Booking.query.filter_by(student_id=student_id).delete(synchronize_session=False)
    InviteToken.query.filter_by(student_id=student_id).delete(synchronize_session=False)
    EmailDispatch.query.filter_by(student_id=student_id).delete(synchronize_session=False)
    Guardian.query.filter_by(student_id=student_id).delete(synchronize_session=False)
    StudentDocument.query.filter_by(student_id=student_id).delete(synchronize_session=False)
    StudentDocumentFolder.query.filter_by(student_id=student_id).delete(synchronize_session=False)
    StudentCV.query.filter_by(student_id=student_id).delete(synchronize_session=False)
    StudentAuth.query.filter_by(student_id=student_id).delete(synchronize_session=False)
    AuditLog.query.filter_by(student_id=student_id).delete(synchronize_session=False)


@students_bp.route("/")
@login_required
def list_students():
    q = request.args.get("q", "").strip()
    branch_filter = request.args.get("branch_id", type=int) or 0
    abroad_only = request.args.get("abroad") == "1"
    page = max(request.args.get("page", type=int) or 1, 1)
    per_page = 20
    query = scope_query_by_branch(Student.query, Student)
    role = normalized_role(current_user.role)
    if role == "IT":
        branch_filter_options = Branch.query.order_by(Branch.name.asc()).all()
    else:
        allowed_ids = user_branch_ids(current_user)
        if allowed_ids:
            branch_filter_options = Branch.query.filter(Branch.id.in_(allowed_ids)).order_by(Branch.name.asc()).all()
        else:
            branch_filter_options = []

    allowed_branch_ids = {b.id for b in branch_filter_options}
    if branch_filter and branch_filter in allowed_branch_ids:
        query = query.filter(Student.branch_id == branch_filter)
    else:
        branch_filter = 0

    if abroad_only:
        query = query.join(StudyCase, StudyCase.student_id == Student.id).filter(
            StudyCase.is_active.is_(True),
            StudyCase.status.in_(["parti", "arrive", "installe"]),
            StudyCase.created_at >= Student.created_at,
        ).distinct()
    else:
        # Exclure de la liste principale les etudiants deja a l'etranger.
        abroad_exists = (
            db.session.query(StudyCase.id)
            .filter(
                StudyCase.student_id == Student.id,
                StudyCase.is_active.is_(True),
                StudyCase.status.in_(["parti", "arrive", "installe"]),
                StudyCase.created_at >= Student.created_at,
            )
            .exists()
        )
        query = query.filter(
            ~abroad_exists
        )

    if q:
        query = query.filter(
            (Student.matricule.ilike(f"%{q}%"))
            | (Student.nom.ilike(f"%{q}%"))
            | (Student.prenoms.ilike(f"%{q}%"))
            | (Student.filiere.ilike(f"%{q}%"))
        )
    pagination = query.order_by(Student.nom.asc(), Student.prenoms.asc(), Student.matricule.asc()).paginate(page=page, per_page=per_page, error_out=False)
    students = pagination.items
    branches = []
    unassigned_count = 0
    if role == "IT":
        branches = Branch.query.order_by(Branch.name.asc()).all()
        unassigned_count = Student.query.filter(Student.branch_id.is_(None), Student.deleted_at.is_(None)).count()
    else:
        allowed_ids = user_branch_ids(current_user)
        if allowed_ids:
            branches = Branch.query.filter(Branch.id.in_(allowed_ids)).order_by(Branch.name.asc()).all()
            unassigned_count = Student.query.filter(Student.branch_id.is_(None), Student.deleted_at.is_(None)).count()

    return render_template(
        "students/list.html",
        students=students,
        pagination=pagination,
        q=q,
        branch_filter=branch_filter,
        branch_filter_options=branch_filter_options,
        abroad_only=abroad_only,
        branches=branches,
        unassigned_count=unassigned_count,
    )


@students_bp.route("/abroad")
@login_required
def list_abroad_students():
    page = max(request.args.get("page", type=int) or 1, 1)
    per_page = 20
    query = (
        scope_query_by_branch(Student.query, Student)
        .join(StudyCase, StudyCase.student_id == Student.id)
        .filter(
            StudyCase.is_active.is_(True),
            StudyCase.status.in_(["parti", "arrive", "installe"]),
            StudyCase.created_at >= Student.created_at,
        )
        .distinct()
    )
    pagination = query.order_by(Student.nom.asc(), Student.prenoms.asc()).paginate(page=page, per_page=per_page, error_out=False)
    students = pagination.items
    return render_template("students/abroad.html", students=students, pagination=pagination)


@students_bp.route("/assign-branch", methods=["POST"])
@login_required
@role_required("FOUNDER", "ADMIN_BRANCH", "EMPLOYEE", "IT")
def assign_unassigned_students_branch():
    branch_id = request.form.get("branch_id", type=int)
    if not branch_id:
        flash("Selectionne une branche.", "danger")
        return redirect(url_for("students.list_students"))

    branch = Branch.query.get(branch_id)
    if not branch:
        flash("Branche introuvable.", "danger")
        return redirect(url_for("students.list_students"))

    if not is_founder() and current_user.branch_id != branch.id:
        flash("Acces refuse pour cette branche.", "danger")
        return redirect(url_for("students.list_students"))

    updated = Student.query.filter(Student.branch_id.is_(None), Student.deleted_at.is_(None)).update({"branch_id": branch.id}, synchronize_session=False)
    db.session.commit()
    add_audit_log(current_user.id, "student_branch_assign_bulk", f"{updated} etudiants sans branche affectes a {branch.name}", branch_id=branch.id, action="student_branch_assign_bulk")
    flash(f"{updated} etudiant(s) sans branche affecte(s) a {branch.name}.", "success")
    return redirect(url_for("students.list_students"))


@students_bp.route("/new", methods=["GET", "POST"])
@login_required
@role_required("FOUNDER", "ADMIN_BRANCH", "EMPLOYEE")
def create_student():
    form = StudentForm()
    set_branch_choices(form)
    if request.method == "GET":
        form.matricule.data = generate_matricule()

    if form.validate_on_submit():
        photo_name = None
        if form.photo.data:
            try:
                photo_name = save_uploaded_file(form.photo.data, current_app.config["PHOTO_UPLOAD_DIR"], current_app.config["ALLOWED_IMAGE_EXTENSIONS"])
            except ValueError as exc:
                flash(str(exc), "danger")
                return render_template("students/form.html", form=form, mode="create")

        branch_id = resolve_actor_branch_id() or form.branch_id.data
        if not branch_id:
            flash("Branche introuvable pour ce compte. Contacte IT.", "danger")
            return render_template("students/form.html", form=form, mode="create")

        student = Student(
            branch_id=branch_id,
            matricule=generate_matricule(),
            nom=form.nom.data.strip(),
            prenoms=form.prenoms.data.strip(),
            sexe=form.sexe.data,
            date_naissance=form.date_naissance.data,
            email=(form.email.data or "").strip().lower() or None,
            procedure_email=(form.procedure_email.data or "").strip().lower() or None,
            procedure_email_password=(form.procedure_email_password.data or "").strip() or None,
            telephone=(form.telephone.data or "").strip() or None,
            adresse=(form.adresse.data or "").strip() or None,
            filiere=form.filiere.data.strip(),
            niveau=form.niveau.data.strip(),
            promotion=form.promotion.data.strip(),
            statut=form.statut.data,
            photo_path=photo_name,
        )
        db.session.add(student)
        db.session.commit()

        # Protection anti-recyclage d'ID SQLite: nettoie toute ancienne donnee liee au meme student_id.
        _purge_student_related_data(student.id)
        db.session.commit()

        # Create student portal credentials (temporary password, must change on first login)
        temp_password = f"Temp{student.id:04d}IF"
        auth = StudentAuth(
            student_id=student.id,
            password_hash=password_hasher.hash(temp_password),
            must_change_password=True,
        )
        db.session.add(auth)
        db.session.commit()

        add_audit_log(current_user.id, "student_create", f"Etudiant {student.matricule} cree", student.id, branch_id=student.branch_id, action="student_create")
        flash(f"Etudiant cree. Portail etudiant: matricule={student.matricule}, mot de passe temporaire={temp_password}", "success")
        return redirect(url_for("students.view_student", student_id=student.id))
    return render_template("students/form.html", form=form, mode="create")


@students_bp.route("/<int:student_id>")
@login_required
def view_student(student_id):
    student = get_active_student_or_404(student_id)
    enforce_student_access(student)
    active_case = StudyCase.query.filter_by(student_id=student.id, is_active=True).order_by(StudyCase.id.desc()).first()

    guardians = Guardian.query.filter_by(student_id=student.id).all()
    folder_filter = normalize_folder_name(request.args.get("folder", ""))
    type_filter = request.args.get("doc_type", "").strip()

    folder_rows = StudentDocumentFolder.query.filter_by(student_id=student.id).order_by(StudentDocumentFolder.created_at.asc()).all()
    folder_names = [normalize_folder_name(f.folder_name) for f in folder_rows]
    folder_names = [f for f in folder_names if f]
    legacy_folders = [normalize_folder_name(x[0]) for x in StudentDocument.query.with_entities(StudentDocument.target_folder).filter_by(student_id=student.id).distinct().all()]
    folder_names = list(dict.fromkeys(folder_names + [f for f in legacy_folders if f]))

    if folder_filter and folder_filter not in folder_names:
        folder_filter = ""

    doc_query = StudentDocument.query.filter_by(student_id=student.id)
    if folder_filter:
        doc_query = doc_query.filter(StudentDocument.target_folder.in_(CANONICAL_TO_STORED.get(folder_filter, [folder_filter])))
    if type_filter:
        doc_query = doc_query.filter(StudentDocument.document_type == type_filter)

    documents = doc_query.order_by(StudentDocument.created_at.desc()).all()
    type_options = [x[0] for x in StudentDocumentForm.document_type.kwargs["choices"]]
    doc_form = StudentDocumentForm()
    folder_form = StudentFolderCreateForm()
    doc_form.folder.data = folder_filter
    folder_cards = [{"key": key, "label": FOLDER_LABELS.get(key, key.upper())} for key in folder_names]
    return render_template(
        "students/view.html",
        student=student,
        active_case=active_case,
        guardians=guardians,
        documents=documents,
        doc_form=doc_form,
        folder_form=folder_form,
        folder_filter=folder_filter,
        type_filter=type_filter,
        folder_cards=folder_cards,
        folder_labels=FOLDER_LABELS,
        type_options=type_options,
    )


@students_bp.route("/<int:student_id>/folders/create", methods=["POST"])
@login_required
@role_required("FOUNDER", "ADMIN_BRANCH", "EMPLOYEE")
def create_folder(student_id):
    student = get_active_student_or_404(student_id)
    enforce_student_access(student)
    form = StudentFolderCreateForm()
    if not form.validate_on_submit():
        flash("Saisis un nom de dossier valide.", "danger")
        return redirect(url_for("students.view_student", student_id=student.id))

    folder_name = normalize_folder_name(form.folder_name.data)
    if not folder_name:
        flash("Nom de dossier invalide.", "danger")
        return redirect(url_for("students.view_student", student_id=student.id))

    exists = StudentDocumentFolder.query.filter_by(student_id=student.id, folder_name=folder_name).first()
    if exists:
        flash("Ce dossier existe deja.", "warning")
        return redirect(url_for("students.view_student", student_id=student.id, folder=folder_name))

    row = StudentDocumentFolder(student_id=student.id, folder_name=folder_name, created_by=current_user.id)
    db.session.add(row)
    db.session.commit()
    add_audit_log(current_user.id, "student_folder_create", f"Dossier {folder_name} cree pour {student.matricule}", student.id, branch_id=student.branch_id, action="folder_create")
    flash("Dossier cree.", "success")
    return redirect(url_for("students.view_student", student_id=student.id, folder=folder_name))


@students_bp.route("/<int:student_id>/cv", methods=["GET", "POST"])
@login_required
@role_required("FOUNDER", "ADMIN_BRANCH", "EMPLOYEE")
def edit_student_cv(student_id):
    student = get_active_student_or_404(student_id)
    enforce_student_access(student)
    cv = get_or_create_student_cv(student.id)
    form = StudentCVForm(obj=cv)

    if form.validate_on_submit():
        cv.profile_text = (form.profile_text.data or "").strip() or None
        cv.contact_details = (form.contact_details.data or "").strip() or None
        cv.hobbies = (form.hobbies.data or "").strip() or None
        cv.languages = (form.languages.data or "").strip() or None
        cv.skills = (form.skills.data or "").strip() or None
        cv.education = (form.education.data or "").strip() or None
        cv.social_links = (form.social_links.data or "").strip() or None
        cv.professional_experience = (form.professional_experience.data or "").strip() or None
        cv.extra_experience = (form.extra_experience.data or "").strip() or None
        cv.software = (form.software.data or "").strip() or None
        cv.show_hobbies = bool(form.show_hobbies.data)
        cv.show_languages = bool(form.show_languages.data)
        cv.show_skills = bool(form.show_skills.data)
        cv.show_education = bool(form.show_education.data)
        cv.show_professional_experience = bool(form.show_professional_experience.data)
        cv.show_extra_experience = bool(form.show_extra_experience.data)
        cv.show_software = bool(form.show_software.data)
        cv.show_social_links = bool(form.show_social_links.data)
        cv.updated_by_user_id = current_user.id
        db.session.commit()
        add_audit_log(
            current_user.id,
            "student_cv_update",
            f"CV mis a jour pour {student.matricule}",
            student_id=student.id,
            branch_id=student.branch_id,
            action="student_cv_update",
        )
        flash("CV enregistre.", "success")
        return redirect(url_for("students.preview_student_cv", student_id=student.id))

    return render_template("students/cv_form.html", student=student, form=form)


@students_bp.route("/<int:student_id>/cv/preview")
@login_required
def preview_student_cv(student_id):
    student = get_active_student_or_404(student_id)
    enforce_student_access(student)
    cv = get_or_create_student_cv(student.id)

    return render_template(
        "students/cv_preview.html",
        student=student,
        cv=cv,
        hobbies_items=_split_multiline(cv.hobbies),
        languages_items=_split_multiline(cv.languages),
        skills_items=_split_multiline(cv.skills),
        education_items=_split_multiline(cv.education),
        social_items=_split_multiline(cv.social_links),
        professional_items=_split_multiline(cv.professional_experience),
        extra_items=_split_multiline(cv.extra_experience),
        software_items=_split_multiline(cv.software),
    )


@students_bp.route("/<int:student_id>/cv/download.pdf")
@login_required
def download_student_cv_pdf(student_id):
    student = get_active_student_or_404(student_id)
    enforce_student_access(student)
    cv = get_or_create_student_cv(student.id)

    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.utils import ImageReader
        from reportlab.pdfgen import canvas
    except ImportError:
        flash("Module PDF manquant: installez reportlab.", "danger")
        return redirect(url_for("students.edit_student_cv", student_id=student.id))

    output = io.BytesIO()
    page_w, page_h = A4
    pdf = canvas.Canvas(output, pagesize=A4)

    top_h = 132
    contact_h = 26
    accent_h = 10
    main_h = page_h - top_h - contact_h - accent_h
    left_w = page_w * 0.31
    right_w = page_w - left_w

    def wrap_text(text, font_name, font_size, max_width):
        words = (text or "").split()
        if not words:
            return []
        lines = []
        current = words[0]
        for w in words[1:]:
            probe = f"{current} {w}"
            if pdf.stringWidth(probe, font_name, font_size) <= max_width:
                current = probe
            else:
                lines.append(current)
                current = w
        lines.append(current)
        return lines

    def draw_section(x, y_top, width, height, title, items, body_text=None):
        pad = 6
        y = y_top - pad
        pdf.setFont("Helvetica-Bold", 13.2)
        pdf.setFillColor(colors.HexColor("#111827"))
        pdf.drawString(x + pad, y, title.upper())
        y -= 18
        pdf.setStrokeColor(colors.HexColor("#9CA3AF"))
        pdf.setLineWidth(0.5)
        pdf.line(x + pad, y, x + width - pad, y)
        y -= 15

        if body_text:
            pdf.setFont("Helvetica", 11.4)
            for para in body_text.split("\n"):
                for line in wrap_text(para.strip(), "Helvetica", 11.4, width - (2 * pad)):
                    if y < y_top - height + 8:
                        return
                    pdf.drawString(x + pad, y, line)
                    y -= 14.2
            y -= 6

        pdf.setFont("Helvetica", 11.2)
        for item in items:
            if y < y_top - height + 8:
                return
            line_text = f"- {item}"
            wrapped = wrap_text(line_text, "Helvetica", 11.2, width - (2 * pad))
            for line in wrapped:
                if y < y_top - height + 8:
                    return
                pdf.drawString(x + pad, y, line)
                y -= 13.8

    # Header background
    pdf.setFillColor(colors.HexColor("#ECEDEE"))
    pdf.rect(0, page_h - top_h, page_w, top_h, fill=1, stroke=0)
    pdf.setFillColor(colors.HexColor("#D9642C"))
    pdf.rect(0, page_h - top_h, 76, top_h, fill=1, stroke=0)

    # Photo
    photo_x = 8
    photo_y = page_h - 124
    photo_w = 108
    photo_h = 108
    if student.photo_path:
        photo_path = os.path.join(current_app.config["PHOTO_UPLOAD_DIR"], student.photo_path)
        if os.path.exists(photo_path):
            try:
                pdf.drawImage(ImageReader(photo_path), photo_x, photo_y, width=photo_w, height=photo_h, preserveAspectRatio=True, mask="auto")
            except Exception:
                pass

    full_name = f"{student.nom} {student.prenoms}".strip()
    name_x = photo_x + photo_w + 14
    pdf.setFillColor(colors.HexColor("#1F2937"))
    pdf.setFont("Helvetica-Bold", 28)
    name_lines = wrap_text(full_name, "Helvetica-Bold", 28, page_w - name_x - 12)
    y_name = page_h - 40
    for ln in name_lines[:2]:
        pdf.drawString(name_x, y_name, ln.upper())
        y_name -= 31

    designation = (student.program_wished or "").strip()
    if designation:
        pdf.setFont("Helvetica", 12)
        pdf.setFillColor(colors.HexColor("#4B5563"))
        pdf.drawString(name_x, page_h - 116, designation)

    # Contact strip
    pdf.setFillColor(colors.HexColor("#D9DBDE"))
    pdf.rect(0, page_h - top_h - contact_h, page_w, contact_h, fill=1, stroke=0)
    pdf.setFillColor(colors.HexColor("#111827"))
    pdf.setFont("Helvetica", 10.8)
    pdf.drawString(8, page_h - top_h - 17, f"Tel: {student.telephone or '-'}")
    pdf.drawString(page_w / 3 + 8, page_h - top_h - 17, f"Email: {student.email or '-'}")
    pdf.drawString((2 * page_w / 3) + 8, page_h - top_h - 17, f"Adresse: {student.adresse or '-'}")

    # Main columns
    main_y = accent_h
    pdf.setFillColor(colors.HexColor("#ECEDEE"))
    pdf.rect(0, main_y, left_w, main_h, fill=1, stroke=0)
    pdf.setFillColor(colors.white)
    pdf.rect(left_w, main_y, right_w, main_h, fill=1, stroke=0)

    # Fixed left regions (fills page even with little content)
    left_top = main_y + main_h - 8
    left_regions = [
        ("Education", _split_multiline(cv.education) or ["Non renseigne."]),
        ("Langues", _split_multiline(cv.languages) or ["Non renseigne."]),
        ("Logiciels", _split_multiline(cv.software) or ["Non renseigne."]),
        ("Centres d'interet", _split_multiline(cv.hobbies) or ["Non renseigne."]),
    ]
    left_region_h = (main_h - 8) / 4.0
    for idx, (title, items) in enumerate(left_regions):
        section_top = left_top - (idx * left_region_h)
        draw_section(0, section_top, left_w, left_region_h, title, items)

    # Right fixed regions
    right_x = left_w
    right_top = main_y + main_h - 8
    h_profile = main_h * 0.24
    h_exp = main_h * 0.38
    h_extra = main_h * 0.20
    h_bottom = main_h - h_profile - h_exp - h_extra

    draw_section(right_x, right_top, right_w, h_profile, "Profil", [], body_text=(cv.profile_text or "Non renseigne."))

    exp_items = []
    if cv.show_professional_experience:
        exp_items.extend(_split_multiline(cv.professional_experience))
    draw_section(right_x, right_top - h_profile, right_w, h_exp, "Experiences professionnelles", exp_items or ["Aucune experience renseignee."])

    extra_items = []
    if cv.show_extra_experience:
        extra_items.extend(_split_multiline(cv.extra_experience))
    draw_section(right_x, right_top - h_profile - h_exp, right_w, h_extra, "Experiences extra-professionnelles", extra_items or ["Aucune experience renseignee."])

    bottom_top = right_top - h_profile - h_exp - h_extra
    draw_section(right_x, bottom_top, right_w / 2, h_bottom, "Expertise", _split_multiline(cv.skills) or ["Non renseigne."])
    social_items = _split_multiline(cv.social_links) or ["Non renseigne."]
    if cv.show_social_links:
        draw_section(right_x + (right_w / 2), bottom_top, right_w / 2, h_bottom, "Suivez-moi", social_items)

    # Accent bar always at bottom
    pdf.setFillColor(colors.HexColor("#1DA0E0"))
    pdf.rect(0, 0, page_w, accent_h, fill=1, stroke=0)

    pdf.showPage()
    pdf.save()
    output.seek(0)
    pdf_name = f"CV_{student.matricule}.pdf"
    add_audit_log(current_user.id, "student_cv_download", f"CV PDF telecharge pour {student.matricule}", student_id=student.id, branch_id=student.branch_id, action="student_cv_download")
    return send_file(output, as_attachment=True, download_name=pdf_name, mimetype="application/pdf")


@students_bp.route("/<int:student_id>/documents/upload", methods=["POST"])
@login_required
@role_required("FOUNDER", "ADMIN_BRANCH", "EMPLOYEE")
def upload_document(student_id):
    student = get_active_student_or_404(student_id)
    enforce_student_access(student)

    form = StudentDocumentForm()
    if not form.validate_on_submit():
        flash("Upload invalide. Verifie le dossier actif, type et fichier.", "danger")
        return redirect(url_for("students.view_student", student_id=student.id))

    file_obj = form.file.data
    if not file_obj or not file_obj.filename:
        flash("Fichier requis.", "danger")
        return redirect(url_for("students.view_student", student_id=student.id))

    target_folder = normalize_folder_name(form.folder.data)
    if not target_folder:
        flash("Ouvre d'abord un dossier avant d'uploader.", "danger")
        return redirect(url_for("students.view_student", student_id=student.id))
    folder_exists = StudentDocumentFolder.query.filter_by(student_id=student.id, folder_name=target_folder).first()
    if not folder_exists:
        flash("Ce dossier n'existe pas. Cree-le d'abord.", "danger")
        return redirect(url_for("students.view_student", student_id=student.id))

    original_filename = file_obj.filename
    upload_dir = os.path.join(current_app.config["STUDENT_DOC_UPLOAD_DIR"], str(student.id), target_folder)

    try:
        stored_filename = save_uploaded_file(file_obj, upload_dir, current_app.config["ALLOWED_DOC_EXTENSIONS"])
    except ValueError as exc:
        flash(str(exc), "danger")
        return redirect(url_for("students.view_student", student_id=student.id))

    stored_path = os.path.join(upload_dir, stored_filename)
    size_bytes = os.path.getsize(stored_path) if os.path.exists(stored_path) else None

    row = StudentDocument(
        student_id=student.id,
        uploaded_by=current_user.id,
        target_folder=target_folder,
        document_type=form.document_type.data,
        original_filename=original_filename,
        stored_filename=stored_filename,
        mime_type=file_obj.mimetype,
        size_bytes=size_bytes,
    )
    db.session.add(row)
    db.session.commit()

    add_audit_log(current_user.id, "student_document_upload", f"Document {row.document_type} charge dans {row.target_folder} pour {student.matricule}", student.id, branch_id=student.branch_id, action="document_upload")
    flash("Document uploadé avec succès.", "success")
    return redirect(url_for("students.view_student", student_id=student.id, folder=target_folder))


@students_bp.route("/<int:student_id>/documents/merge-download", methods=["POST"])
@login_required
@role_required("FOUNDER", "ADMIN_BRANCH", "EMPLOYEE")
def merge_documents(student_id):
    student = get_active_student_or_404(student_id)
    enforce_student_access(student)

    selected_ids = request.form.getlist("doc_ids")
    folder = normalize_folder_name(request.form.get("folder", ""))
    target_mb = request.form.get("target_mb", type=int) or 4
    order_map = {}

    if not selected_ids:
        flash("Selectionne au moins un fichier a fusionner.", "warning")
        return redirect(url_for("students.view_student", student_id=student.id, folder=folder))

    for raw_id in selected_ids:
        try:
            doc_id = int(raw_id)
        except ValueError:
            continue
        order_map[doc_id] = request.form.get(f"order_{doc_id}", type=int) or 9999

    docs = StudentDocument.query.filter(
        StudentDocument.student_id == student.id,
        StudentDocument.id.in_([int(x) for x in selected_ids if str(x).isdigit()]),
    ).all()
    if not docs:
        flash("Aucun document valide selectionne.", "warning")
        return redirect(url_for("students.view_student", student_id=student.id, folder=folder))

    docs = sorted(docs, key=lambda d: (order_map.get(d.id, 9999), d.id))
    profile = _compression_profile(target_mb)

    try:
        from pypdf import PdfReader, PdfWriter
        from PIL import Image
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.utils import ImageReader
        from reportlab.pdfgen import canvas
    except ImportError:
        flash("Module manquant pour fusion/compression. Installez: pypdf pillow", "danger")
        return redirect(url_for("students.view_student", student_id=student.id, folder=folder))

    writer = PdfWriter()
    ignored = []
    added = 0

    def append_pdf(path):
        nonlocal added
        reader = PdfReader(path)
        for page in reader.pages:
            try:
                page.compress_content_streams()
            except Exception:
                pass
            writer.add_page(page)
            added += 1

    def append_image(path):
        nonlocal added
        with Image.open(path) as img:
            img = img.convert("RGB")
            img.thumbnail((profile["max_dim"], profile["max_dim"]))

            with tempfile.NamedTemporaryFile(suffix=".jpg", delete=False) as tmp_img:
                tmp_img_path = tmp_img.name
            img.save(tmp_img_path, format="JPEG", quality=profile["jpeg_quality"], optimize=True)

            with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp_pdf:
                tmp_pdf_path = tmp_pdf.name

            w, h = A4
            c = canvas.Canvas(tmp_pdf_path, pagesize=A4)
            iw, ih = img.size
            scale = min((w - 30) / iw, (h - 30) / ih)
            dw, dh = iw * scale, ih * scale
            x = (w - dw) / 2
            y = (h - dh) / 2
            c.drawImage(ImageReader(tmp_img_path), x, y, width=dw, height=dh, preserveAspectRatio=True, mask="auto")
            c.showPage()
            c.save()

            os.remove(tmp_img_path)
            append_pdf(tmp_pdf_path)
            os.remove(tmp_pdf_path)

    for d in docs:
        base_dir = os.path.join(current_app.config["STUDENT_DOC_UPLOAD_DIR"], str(student.id), d.target_folder)
        path = os.path.join(base_dir, d.stored_filename)
        if not os.path.exists(path):
            ignored.append(f"{d.original_filename} (introuvable)")
            continue
        ext = (d.original_filename.rsplit(".", 1)[1].lower() if "." in d.original_filename else "")
        if ext == "pdf":
            append_pdf(path)
        elif ext in {"jpg", "jpeg", "png", "webp"}:
            append_image(path)
        else:
            ignored.append(f"{d.original_filename} (non fusionnable)")

    if added == 0:
        flash("Aucun fichier fusionnable (PDF ou image).", "warning")
        return redirect(url_for("students.view_student", student_id=student.id, folder=folder))

    output = io.BytesIO()
    writer.write(output)
    output.seek(0)
    size_mb = len(output.getvalue()) / (1024 * 1024)

    msg = f"Fusion terminee: {added} page(s)."
    if ignored:
        msg += f" Ignorés: {len(ignored)}."
    if size_mb > target_mb:
        msg += f" Taille finale {size_mb:.2f} MB (> {target_mb} MB)."
        flash(msg, "warning")
    else:
        flash(msg, "success")

    add_audit_log(current_user.id, "student_document_merge", f"Fusion docs pour {student.matricule}: {added} pages, cible={target_mb}MB", student.id, branch_id=student.branch_id, action="document_merge")
    filename = f"{student.matricule}_documents_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.pdf"
    return send_file(output, as_attachment=True, download_name=filename, mimetype="application/pdf")


@students_bp.route("/<int:student_id>/documents/<int:document_id>/download")
@login_required
def download_document(student_id, document_id):
    student = get_active_student_or_404(student_id)
    enforce_student_access(student)

    document = StudentDocument.query.filter_by(id=document_id, student_id=student.id).first_or_404()
    base_dir = os.path.join(current_app.config["STUDENT_DOC_UPLOAD_DIR"], str(student.id), document.target_folder)
    file_path = os.path.join(base_dir, document.stored_filename)
    if not os.path.exists(file_path):
        flash("Fichier introuvable sur le serveur.", "danger")
        return redirect(url_for("students.view_student", student_id=student.id))

    add_audit_log(current_user.id, "student_document_download", f"Telechargement document #{document.id}", student.id, branch_id=student.branch_id, action="document_download")
    return send_from_directory(base_dir, document.stored_filename, as_attachment=True, download_name=document.original_filename)


@students_bp.route("/<int:student_id>/documents/<int:document_id>/delete", methods=["POST"])
@login_required
@role_required("FOUNDER", "ADMIN_BRANCH", "EMPLOYEE")
def delete_document(student_id, document_id):
    student = get_active_student_or_404(student_id)
    enforce_student_access(student)

    document = StudentDocument.query.filter_by(id=document_id, student_id=student.id).first_or_404()
    base_dir = os.path.join(current_app.config["STUDENT_DOC_UPLOAD_DIR"], str(student.id), document.target_folder)
    file_path = os.path.join(base_dir, document.stored_filename)
    if os.path.exists(file_path):
        os.remove(file_path)

    db.session.delete(document)
    db.session.commit()
    add_audit_log(current_user.id, "student_document_delete", f"Suppression document #{document_id}", student.id, branch_id=student.branch_id, action="document_delete")
    flash("Document supprime.", "success")
    return redirect(url_for("students.view_student", student_id=student.id, folder=normalize_folder_name(document.target_folder)))


@students_bp.route("/<int:student_id>/edit", methods=["GET", "POST"])
@login_required
@role_required("FOUNDER", "ADMIN_BRANCH", "EMPLOYEE")
def edit_student(student_id):
    student = get_active_student_or_404(student_id)
    enforce_student_access(student)

    form = StudentForm(obj=student)
    set_branch_choices(form)
    if request.method == "GET":
        form.branch_id.data = student.branch_id
    if form.validate_on_submit():
        student.nom = form.nom.data.strip()
        student.prenoms = form.prenoms.data.strip()
        student.sexe = form.sexe.data
        student.date_naissance = form.date_naissance.data
        student.email = (form.email.data or "").strip().lower() or None
        student.procedure_email = (form.procedure_email.data or "").strip().lower() or None
        student.procedure_email_password = (form.procedure_email_password.data or "").strip() or None
        student.telephone = (form.telephone.data or "").strip() or None
        student.adresse = (form.adresse.data or "").strip() or None
        student.filiere = form.filiere.data.strip()
        student.niveau = form.niveau.data.strip()
        student.promotion = form.promotion.data.strip()
        student.statut = form.statut.data
        actor_branch_id = resolve_actor_branch_id()
        if actor_branch_id:
            student.branch_id = actor_branch_id

        if form.photo.data:
            try:
                student.photo_path = save_uploaded_file(form.photo.data, current_app.config["PHOTO_UPLOAD_DIR"], current_app.config["ALLOWED_IMAGE_EXTENSIONS"])
            except ValueError as exc:
                flash(str(exc), "danger")
                return render_template("students/form.html", form=form, mode="edit")

        db.session.commit()
        add_audit_log(current_user.id, "student_update", f"Etudiant {student.matricule} modifie", student.id, branch_id=student.branch_id, action="student_update")
        flash("Etudiant modifie.", "success")
        return redirect(url_for("students.view_student", student_id=student.id))
    return render_template("students/form.html", form=form, mode="edit")


@students_bp.route("/<int:student_id>/delete", methods=["POST"])
@login_required
@role_required("FOUNDER", "ADMIN_BRANCH")
def delete_student(student_id):
    student = get_active_student_or_404(student_id)
    enforce_student_access(student)

    matricule = student.matricule
    _purge_student_related_data(student.id)
    student.deleted_at = datetime.utcnow()
    student.statut = "ancien"
    db.session.commit()
    add_audit_log(current_user.id, "student_delete", f"Etudiant {matricule} supprime (soft delete)", branch_id=student.branch_id, action="student_delete")
    flash("Etudiant supprime.", "success")
    return redirect(url_for("students.list_students"))


@students_bp.route("/<int:student_id>/guardians/new", methods=["GET", "POST"])
@login_required
@role_required("FOUNDER", "ADMIN_BRANCH", "EMPLOYEE")
def create_guardian(student_id):
    student = get_active_student_or_404(student_id)
    enforce_student_access(student)

    form = GuardianForm()
    if form.validate_on_submit():
        guardian = Guardian(
            student_id=student.id,
            nom=form.nom.data.strip(),
            prenoms=form.prenoms.data.strip(),
            lien_parente=form.lien_parente.data.strip(),
            telephone=(form.telephone.data or "").strip() or None,
            email=(form.email.data or "").strip().lower() or None,
            adresse=(form.adresse.data or "").strip() or None,
            contact_urgence=form.contact_urgence.data == "1",
        )
        db.session.add(guardian)
        db.session.commit()
        add_audit_log(current_user.id, "guardian_create", f"Parent ajoute pour {student.matricule}", student.id, branch_id=student.branch_id, action="guardian_create")
        flash("Responsable ajoute.", "success")
        return redirect(url_for("students.view_student", student_id=student.id))
    return render_template("students/guardian_form.html", form=form, student=student)


@students_bp.route("/export")
@login_required
def export_students_csv():
    students = scope_query_by_branch(Student.query, Student).order_by(Student.nom.asc(), Student.prenoms.asc(), Student.matricule.asc()).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Etudiants"

    headers = [
        "Matricule",
        "Nom",
        "Prenoms",
        "Sexe",
        "Date Naissance",
        "Email",
        "Telephone",
        "Adresse",
        "Filiere",
        "Niveau",
        "Promotion",
        "Statut",
    ]

    title = f"InnovFormation - Export Etudiants ({datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')})"
    ws.merge_cells(start_row=1, start_column=3, end_row=2, end_column=len(headers))
    ws.cell(row=1, column=3, value=title)
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 24
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=2)
    ws.cell(row=1, column=1, value="INNOVFORMATION")

    title_cell = ws.cell(row=1, column=3)
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor="1E3A8A")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    logo_text = ws.cell(row=1, column=1)
    logo_text.font = Font(bold=True, size=12, color="1E3A8A")
    logo_text.alignment = Alignment(horizontal="center", vertical="center")
    logo_text.fill = PatternFill("solid", fgColor="DBEAFE")

    logo_candidates = [
        os.path.join(current_app.root_path, "static", "partners", "innovformation.png"),
        os.path.join(current_app.root_path, "static", "partners", "logo.png"),
    ]
    for logo_path in logo_candidates:
        if os.path.exists(logo_path):
            try:
                from openpyxl.drawing.image import Image as XLImage

                logo_img = XLImage(logo_path)
                logo_img.width = 170
                logo_img.height = 48
                ws.add_image(logo_img, "A1")
                break
            except Exception:
                break

    header_fill = PatternFill("solid", fgColor="1D4ED8")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_row = 4

    for idx, header in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=idx, value=header)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border
    ws.row_dimensions[header_row].height = 22

    status_colors = {"actif": "DCFCE7", "suspendu": "FEF3C7", "ancien": "E5E7EB"}
    data_start = header_row + 1

    for i, s in enumerate(students, start=data_start):
        values = [
            s.matricule,
            s.nom,
            s.prenoms,
            s.sexe,
            s.date_naissance.isoformat() if s.date_naissance else "",
            s.email or "",
            s.telephone or "",
            s.adresse or "",
            s.filiere,
            s.niveau,
            s.promotion,
            s.statut,
        ]
        for j, v in enumerate(values, start=1):
            c = ws.cell(row=i, column=j, value=v)
            c.border = border
            c.alignment = Alignment(vertical="top", wrap_text=True)
            if i % 2 == 0:
                c.fill = PatternFill("solid", fgColor="F8FAFC")

        status_cell = ws.cell(row=i, column=12)
        status_fill = status_colors.get((s.statut or "").lower())
        if status_fill:
            status_cell.fill = PatternFill("solid", fgColor=status_fill)
        status_cell.alignment = Alignment(horizontal="center", vertical="center")

    if not students:
        for j in range(1, len(headers) + 1):
            c = ws.cell(row=data_start, column=j, value="")
            c.border = border

    last_row = max(data_start, data_start + len(students) - 1)
    ws.freeze_panes = f"A{data_start}"

    col_widths = {
        "A": 20,
        "B": 18,
        "C": 22,
        "D": 8,
        "E": 16,
        "F": 30,
        "G": 16,
        "H": 34,
        "I": 20,
        "J": 14,
        "K": 12,
        "L": 12,
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    table = Table(displayName="StudentsTable", ref=f"A{header_row}:L{last_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)

    stats_ws = wb.create_sheet("Statistiques")
    stats_ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    stats_ws.cell(row=1, column=1, value="Statistiques export etudiants")
    stats_ws.cell(row=1, column=1).font = Font(bold=True, size=13, color="FFFFFF")
    stats_ws.cell(row=1, column=1).fill = PatternFill("solid", fgColor="1E3A8A")
    stats_ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
    stats_ws.row_dimensions[1].height = 24

    def write_counter_block(start_row, title_text, counter_obj):
        stats_ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=3)
        t = stats_ws.cell(row=start_row, column=1, value=title_text)
        t.font = Font(bold=True, color="1E3A8A")
        t.fill = PatternFill("solid", fgColor="DBEAFE")
        t.alignment = Alignment(horizontal="left", vertical="center")
        h1 = stats_ws.cell(row=start_row + 1, column=1, value="Valeur")
        h2 = stats_ws.cell(row=start_row + 1, column=2, value="Total")
        for h in (h1, h2):
            h.font = Font(bold=True, color="FFFFFF")
            h.fill = PatternFill("solid", fgColor="1D4ED8")
            h.alignment = Alignment(horizontal="center", vertical="center")
            h.border = border
        row = start_row + 2
        for key, value in counter_obj.items():
            c1 = stats_ws.cell(row=row, column=1, value=key)
            c2 = stats_ws.cell(row=row, column=2, value=value)
            c1.border = border
            c2.border = border
            row += 1
        if row == start_row + 2:
            c1 = stats_ws.cell(row=row, column=1, value="N/A")
            c2 = stats_ws.cell(row=row, column=2, value=0)
            c1.border = border
            c2.border = border
        return row + 1

    filiere_counter = Counter((s.filiere or "N/A") for s in students)
    niveau_counter = Counter((s.niveau or "N/A") for s in students)
    promotion_counter = Counter((s.promotion or "N/A") for s in students)

    next_row = 3
    next_row = write_counter_block(next_row, "Totaux par Filiere", filiere_counter)
    next_row = write_counter_block(next_row, "Totaux par Niveau", niveau_counter)
    write_counter_block(next_row, "Totaux par Promotion", promotion_counter)

    stats_ws.column_dimensions["A"].width = 34
    stats_ws.column_dimensions["B"].width = 14

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"innovformation_etudiants_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@students_bp.route("/import", methods=["GET", "POST"])
@csrf.exempt
@login_required
@role_required("FOUNDER", "ADMIN_BRANCH", "EMPLOYEE", "IT")
def import_students_csv():
    if request.method == "POST":
        uploaded = request.files.get("file")
        if not uploaded:
            flash("Fichier requis.", "danger")
            return redirect(url_for("students.import_students_csv"))

        raw_content = uploaded.read()
        if not raw_content:
            flash("Fichier vide ou illisible.", "danger")
            return redirect(url_for("students.import_students_csv"))

        filename = (uploaded.filename or "").strip().lower()

        created = 0
        updated = 0
        skipped = 0
        reserved_matricules = set()
        target_branch_id = resolve_actor_branch_id()
        if not target_branch_id:
            flash("Branche introuvable pour cet import. Connecte-toi avec un compte agence ou choisis un scope IT.", "danger")
            return redirect(url_for("students.import_students_csv"))
        key_aliases = {
            "matricule": "matricule",
            "client_id": "matricule",
            "code_client": "matricule",
            "nom": "nom",
            "prenoms": "prenoms",
            "prenom": "prenoms",
            "sexe": "sexe",
            "date_naissance": "date_naissance",
            "datenaissance": "date_naissance",
            "date_de_naissance": "date_naissance",
            "email": "email",
            "telephone": "telephone",
            "tel": "telephone",
            "contact_eleve": "telephone",
            "contact_parent": "telephone",
            "adresse": "adresse",
            "filiere": "filiere",
            "programmes_filieres": "filiere",
            "programme_filiere": "filiere",
            "niveau": "niveau",
            "promotion": "promotion",
            "annee": "promotion",
            "annee_scolaire": "promotion",
            "year": "promotion",
            "last_name": "nom",
            "firstname": "prenoms",
            "first_name": "prenoms",
            "surname": "nom",
            "phone": "telephone",
            "mobile": "telephone",
            "mail": "email",
            "e_mail": "email",
            "class": "niveau",
            "niveau_etude": "niveau",
            "program": "filiere",
            "programme": "filiere",
            "statut": "statut",
            "statut_global": "statut_global",
        }

        def _norm_key(raw_key):
            base = (raw_key or "").strip().lower()
            base = unicodedata.normalize("NFKD", base)
            base = "".join(ch for ch in base if not unicodedata.combining(ch))
            base = re.sub(r"[^a-z0-9]+", "_", base).strip("_")
            return key_aliases.get(base, "")

        def _guess_niveau(filiere_text):
            txt = (filiere_text or "").strip().upper()
            m = re.search(r"(BTS\s*\d+|LICENCE\s*\d+|MASTER\s*\d+|DOCTORAT\s*\d+|MBA\s*\d+|DUT\s*\d+|BACHELOR\s*\d+)", txt)
            if m:
                return m.group(1)
            return ""

        def _is_internal_matricule(value):
            v = (value or "").strip().upper()
            return bool(re.fullmatch(r"IF-\d{4}-\d{5}", v))

        row_source = []
        if filename.endswith(".xls"):
            flash("Format .xls non supporte. Convertis d'abord en .xlsx puis reimporte.", "danger")
            return redirect(url_for("students.import_students_csv"))

        if filename.endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
            try:
                wb = load_workbook(io.BytesIO(raw_content), read_only=True, data_only=True)
            except Exception:
                flash("Fichier Excel invalide ou illisible.", "danger")
                return redirect(url_for("students.import_students_csv"))

            # Find the best sheet/header candidate across workbook.
            best_sheet = None
            best_header = None
            best_header_row = None
            best_score = -1

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for row_idx, r in enumerate(ws.iter_rows(min_row=1, max_row=40, values_only=True), start=1):
                    vals = [("" if v is None else str(v)).strip() for v in r]
                    if not any(vals):
                        continue
                    mapped = [_norm_key(v) for v in vals]
                    mapped_nonempty = [m for m in mapped if m]
                    score = len(set(mapped_nonempty))
                    has_identity = ("nom" in mapped_nonempty and "prenoms" in mapped_nonempty)
                    has_program = ("filiere" in mapped_nonempty or "niveau" in mapped_nonempty or "promotion" in mapped_nonempty)
                    if score > best_score and (has_identity or has_program):
                        best_score = score
                        best_sheet = ws
                        best_header = vals
                        best_header_row = row_idx

            if not best_sheet or not best_header or best_score < 2:
                flash("Excel invalide: impossible de detecter une ligne d'en-tetes compatible.", "danger")
                return redirect(url_for("students.import_students_csv"))

            for r in best_sheet.iter_rows(min_row=(best_header_row + 1), values_only=True):
                vals = [("" if v is None else str(v)).strip() for v in r]
                if not any(vals):
                    continue
                row = {}
                for idx, hv in enumerate(best_header):
                    nk = _norm_key(hv)
                    if nk:
                        row[nk] = vals[idx] if idx < len(vals) else ""
                if row:
                    row_source.append(row)
        elif filename.endswith(".pdf"):
            try:
                from pypdf import PdfReader
            except Exception:
                flash("Import PDF indisponible: installez pypdf.", "danger")
                return redirect(url_for("students.import_students_csv"))

            try:
                reader = PdfReader(io.BytesIO(raw_content))
                full_text = "\n".join((page.extract_text() or "") for page in reader.pages)
            except Exception:
                flash("PDF invalide ou illisible.", "danger")
                return redirect(url_for("students.import_students_csv"))

            lines = [ln.strip() for ln in full_text.splitlines() if ln and ln.strip()]
            if not lines:
                flash("PDF vide ou sans texte exploitable.", "danger")
                return redirect(url_for("students.import_students_csv"))

            def _split_pdf_line(line, delimiter):
                if delimiter == "whitespace":
                    return [x.strip() for x in re.split(r"\s{2,}", line) if x and x.strip()]
                return [x.strip() for x in line.split(delimiter)]

            best_idx = None
            best_delim = None
            best_tokens = None
            best_score = -1
            delimiters = [";", ",", "|", "	", "whitespace"]

            for idx, line in enumerate(lines[:120]):
                for delim in delimiters:
                    tokens = _split_pdf_line(line, delim)
                    if len(tokens) < 2:
                        continue
                    mapped = [_norm_key(tok) for tok in tokens]
                    mapped_nonempty = [m for m in mapped if m]
                    score = len(set(mapped_nonempty))
                    has_identity = ("nom" in mapped_nonempty and "prenoms" in mapped_nonempty)
                    has_program = ("filiere" in mapped_nonempty or "niveau" in mapped_nonempty or "promotion" in mapped_nonempty)
                    if score > best_score and (has_identity or has_program):
                        best_score = score
                        best_idx = idx
                        best_delim = delim
                        best_tokens = tokens

            if best_idx is None or best_tokens is None or best_score < 2:
                flash("PDF invalide: impossible de detecter une ligne d'en-tetes compatible.", "danger")
                return redirect(url_for("students.import_students_csv"))

            for line in lines[best_idx + 1:]:
                vals = _split_pdf_line(line, best_delim)
                if len(vals) < 2:
                    continue
                row = {}
                for idx, hv in enumerate(best_tokens):
                    nk = _norm_key(hv)
                    if nk:
                        row[nk] = vals[idx] if idx < len(vals) else ""
                if row:
                    row_source.append(row)
        else:
            content = raw_content.decode("utf-8-sig", errors="replace")
            sample = content[:4096]
            lines = content.splitlines()
            first_line = lines[0] if lines else ""
            forced_delim = None
            delim_scores = {
                ";": first_line.count(";"),
                ",": first_line.count(","),
                "	": first_line.count("	"),
                "|": first_line.count("|"),
            }
            best_delim = max(delim_scores, key=delim_scores.get) if first_line else None
            if best_delim and delim_scores.get(best_delim, 0) > 0:
                forced_delim = best_delim

            if forced_delim:
                reader = csv.DictReader(io.StringIO(content, newline=""), delimiter=forced_delim)
            else:
                try:
                    dialect = csv.Sniffer().sniff(sample, delimiters=",;|	")
                except csv.Error:
                    dialect = csv.excel
                reader = csv.DictReader(io.StringIO(content, newline=""), dialect=dialect)

            if not reader.fieldnames:
                flash("CSV invalide: en-tetes manquants.", "danger")
                return redirect(url_for("students.import_students_csv"))

            if len(reader.fieldnames) == 1 and first_line:
                retry_delim = None
                for d in [";", ",", "	", "|"]:
                    if first_line.count(d) > 0:
                        retry_delim = d
                        break
                if retry_delim:
                    reader = csv.DictReader(io.StringIO(content, newline=""), delimiter=retry_delim)

            for raw_row in reader:
                row = {}
                for k, v in raw_row.items():
                    nk = _norm_key(k)
                    if nk:
                        row[nk] = (v or "")
                if row:
                    row_source.append(row)

        try:
            for row in row_source:
                normalized = {k: (row.get(k) or "").strip() for k in (
                    "matricule", "nom", "prenoms", "sexe", "date_naissance", "email",
                    "telephone", "adresse", "filiere", "niveau", "promotion", "statut", "statut_global"
                )}

                # Ignore fully empty lines from CSV files.
                if not any(normalized.values()):
                    skipped += 1
                    continue

                raw_matricule = normalized["matricule"]
                matricule = raw_matricule.upper() if _is_internal_matricule(raw_matricule) else ""
                if matricule and matricule in reserved_matricules:
                    skipped += 1
                    continue

                date_naissance = None
                if normalized["date_naissance"]:
                    try:
                        date_naissance = datetime.fromisoformat(normalized["date_naissance"]).date()
                    except ValueError:
                        skipped += 1
                        continue

                nom_val = normalized["nom"]
                prenoms_val = normalized["prenoms"]
                filiere_val = normalized["filiere"]
                niveau_val = normalized["niveau"] or _guess_niveau(normalized["filiere"])
                promotion_val = normalized["promotion"]
                sexe_val = normalized["sexe"] or "M"
                statut_val = normalized["statut"] or "actif"
                statut_global_val = normalized["statut_global"] or "prospect"
                email_val = normalized["email"].lower() if normalized["email"] else None
                telephone_val = normalized["telephone"] or None
                adresse_val = normalized["adresse"] or None

                existing = Student.query.filter(Student.matricule == matricule).first() if matricule else None
                if existing is None and email_val:
                    existing = Student.query.filter(Student.email == email_val, Student.deleted_at.is_(None)).order_by(Student.id.asc()).first()
                if existing is None and nom_val and prenoms_val and promotion_val:
                    existing = (
                        Student.query.filter(
                            Student.nom == nom_val,
                            Student.prenoms == prenoms_val,
                            Student.promotion == promotion_val,
                            Student.deleted_at.is_(None),
                        )
                        .order_by(Student.id.asc())
                        .first()
                    )

                if existing:
                    existing.branch_id = target_branch_id
                    if nom_val:
                        existing.nom = nom_val
                    if prenoms_val:
                        existing.prenoms = prenoms_val
                    if sexe_val:
                        existing.sexe = sexe_val
                    existing.date_naissance = date_naissance
                    existing.email = email_val
                    existing.telephone = telephone_val
                    existing.adresse = adresse_val
                    if filiere_val:
                        existing.filiere = filiere_val
                    if niveau_val:
                        existing.niveau = niveau_val
                    if promotion_val:
                        existing.promotion = promotion_val
                    if statut_val:
                        existing.statut = statut_val
                    if statut_global_val:
                        existing.statut_global = statut_global_val
                    existing.deleted_at = None
                    updated += 1
                else:
                    # For a new student, required fields must be provided.
                    if not (nom_val and prenoms_val and filiere_val and niveau_val and promotion_val):
                        skipped += 1
                        continue

                    if not matricule:
                        matricule = generate_matricule(reserved=reserved_matricules)
                        reserved_matricules.add(matricule)

                    student = Student(
                        branch_id=target_branch_id,
                        matricule=matricule,
                        nom=nom_val,
                        prenoms=prenoms_val,
                        sexe=sexe_val,
                        date_naissance=date_naissance,
                        email=email_val,
                        telephone=telephone_val,
                        adresse=adresse_val,
                        filiere=filiere_val,
                        niveau=niveau_val,
                        promotion=promotion_val,
                        statut=statut_val,
                        statut_global=statut_global_val,
                    )
                    db.session.add(student)
                    created += 1

                if matricule:
                    reserved_matricules.add(matricule)
        except csv.Error as exc:
            db.session.rollback()
            flash(f"CSV invalide: {exc}", "danger")
            return redirect(url_for("students.import_students_csv"))

        db.session.commit()
        add_audit_log(current_user.id, "student_import", f"Import fichier: {created} crees, {updated} mis a jour, {skipped} ignores", branch_id=current_user.branch_id, action="student_import")
        flash(f"Import termine. Crees: {created}, mis a jour: {updated}, ignores: {skipped}", "success")
        return redirect(url_for("students.list_students"))

    return render_template("students/import.html")












